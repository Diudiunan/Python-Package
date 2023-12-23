from anyio import sleep
import pandas as pd
import csv
import os
import re
import datetime
import openpyxl


excel_file = '再犯严重性裁判文书数据1.xlsx'
#excel_file = 'text.xlsx'


def write_to_column(data_list,ColumnInt):
    # sheet名称
    sheet_name = 'Sheet1'
    # 获取指定的文件
    wb = openpyxl.load_workbook(excel_file)
    # 获取指定的sheet
    ws = wb[sheet_name]
    # 获得最大列数
    max_col_num = ws.max_column
    for index, each in enumerate(data_list):
        # 通过row关键字指定行，colunm关键字指定列，均从1开始
        ws.cell(row=index + 1, column=max_col_num + ColumnInt, value=each)

    # 保存文件
    wb.save(excel_file)


def delete_row(dataList):
    # sheet名称
    sheet_name = 'Sheet1'
    # 获取指定的文件
    wb = openpyxl.load_workbook(excel_file)
    # 获取指定的sheet
    ws = wb[sheet_name]
    # 获得最大列数
    max_col_num = ws.max_column
    for index, each in enumerate(dataList):
        # 通过row关键字指定行，colunm关键字指定列，均从1开始
        ws.cell(row=index + 1, column=max_col_num , value=each)

    # 保存文件
    wb.save(excel_file)


def parse_date_to_ymd(para):
    if para == '.':
        return '.'
    delta = pd.Timedelta(str(para)+'days')
    time = pd.to_datetime('1899-12-30') + delta
    time = re.sub(r' [0-9]{2}:[0-9]{2}:[0-9]{2}$', '', str(time))
    return str(time)


def cal_count_of_day(date_str1, date_str2):
    date1 = datetime.datetime.strptime(date_str1, '%Y-%m-%d').date()
    date2 = datetime.datetime.strptime(date_str2, '%Y-%m-%d').date()
    delta = date2 - date1
    days = delta.days
    return days

#汉字转阿拉伯数字，通用方法
def NumberStrToInt(StrList):
    #一亿内
    NumberMap = {'一':1,'二':2,'三':3,'四':4,'五':5,'六':6,'七':7,'八':8,'九':9}
    UnitMap = {'十':10,'百':100,'千':1000,'万':10000}
    Number = 0
    Sum = 0
    Unit = 1
    BigUnit = 1
    for index, j in enumerate(reversed(StrList)):
        if j in NumberMap:
            Number = NumberMap[j]
            Sum += Number * Unit
        elif j in UnitMap:
            Unit = UnitMap[j]
            if Unit % 10000 == 0:
                BigUnit = Unit
            if Unit < BigUnit:
                Unit *= BigUnit
            if index == len(StrList)-1:
                Sum += 1 * Unit
    return Sum

#汉字时间转天数数字,belong F2
def ChangeToDays(Str):
    NumberYear = 0
    NumberMon = 0
    NumberDay = 0
    Str = str(Str)
    if '.' in Str or '.' == Str:
        return '.'
    if '无期' in Str:
        return 'W'
    if '年' in Str:
        YearIndex = Str.index('年')
        NumberYear = NumberStrToInt(Str[0:YearIndex])
        if '月' in Str:
            NumberMon = NumberStrToInt(Str[YearIndex+1:Str.index('月')-1])
            if '日' in Str:
                NumberDay = NumberStrToInt(Str[Str.index('月'):Str.index('日')])
    else:
        if '月' in Str:
            NumberMon = NumberStrToInt(Str[0:Str.index('月')-1])
            if '日' in Str:
                NumberDay = NumberStrToInt(Str[Str.index('月'):Str.index('日')])
        else:
            if '日' in Str:
                NumberDay = NumberStrToInt(Str[:Str.index('日')])
    ReaultNumber = NumberYear*365 + NumberMon*30 +NumberDay
    if ReaultNumber == 0:
        return '.'
    return ReaultNumber

#汉字金额转数字金额,belong F4
def ChangeToNumber(Str):
    if '元' in Str:
        YuanIndex = Str.index('元')
        if YuanIndex == 0:
            return '.'
        elif Str[YuanIndex-1].isdigit():
            for i,j in enumerate(Str[:YuanIndex]):
                if j.isdigit():
                    return int(Str[i:YuanIndex])
            return '.'
        else:
            return NumberStrToInt(Str[:YuanIndex])
    else:
        return '.'


def to_0_1(para):
    para = str(para)
    if para == '0':
        return para
    elif para == '1':
        return para
    else:
        return '.'


def set_birthday(para):
    para = str(para)
    print(para)

def task1_func():
    data = pd.read_excel(excel_file)
    loc1 = data['时间1']
    loc2 = data['时间2']
    final_time3 = ['时间3']
    for i in range(0, int(loc2.count())):
        try:
            time1 = parse_date_to_ymd(loc1[i])
            time2 = parse_date_to_ymd(loc2[i])
            if time1 == '.':
                final_time3.append('.')
                continue
            if time2 == '.':
                final_time3.append('.')
                continue
            time3 = cal_count_of_day(time1, time2)
            final_time3.append(time3 if time3>=0 else '.')
        except:
            final_time3.append('.')
            continue

        #print(time1, time2, time3)
    write_to_column(final_time3,1)

def task2_func():
    data = pd.read_excel(excel_file)
    loc = data['曾犯罪被判处的监禁时间']
    AnList = ['曾犯罪被判处的监禁时间(天)']
    for i in range(0, int(loc.count())):
        res = ChangeToDays(loc[i])
        AnList.append(res)
    write_to_column(AnList,1)

def task3_func():
    data = pd.read_excel(excel_file)
    loc1 = data['时间3']
    loc2 = data['曾犯罪被判处的监禁时间(天)']
    List = ['实际监禁时间/曾被判监禁时间']
    rate = -1
    for i, j in enumerate(loc1):
        if j == '.' or loc2[i] == '.':
            List.append('.')
            continue
        if isinstance(j,int):
            if isinstance(loc2[i],int) or loc2[i].isdigit():
                rate = j/int(loc2[i])
        elif j.isdigit():
            if isinstance(loc2[i],int) or loc2[i].isdigit():
                rate = int(j)/int(loc2[i])
        if 0< rate <= 1:
            List.append(round(rate,3))
        else:
            List.append('.')
    write_to_column(List,1)


def task4_func():
    data = pd.read_excel(excel_file)
    loc = data['曾犯罪被判处的罚金']
    data_list = ['曾犯罪被判处的罚金(元)']
    for i in range(0, int(loc.count())):
        try:
            res = ChangeToNumber(loc[i])
            data_list.append(res)
        except:
            data_list.append('.')
            continue
    write_to_column(data_list,1)


def task5_func():
    data = pd.read_excel(excel_file)
    loc = data['民族']
    data = ['民族(0汉/1少)']
    for i in loc:
        res = to_0_1(i)
        data.append(res)
    write_to_column(data,1)


def task6_func():
    data = pd.read_excel(excel_file)
    loc = data['出生日期']
    data = ['出生日期(年)']
    for i in loc:
        try:
            date = parse_date_to_ymd(i)
            date = date.split('-')[0]
            data.append(date)
        except:
            data.append('.')
            continue
    write_to_column(data,1)

def task7_func():
    data = pd.read_excel(excel_file)
    loc1 = data['出生日期(年)']
    loc2 = data['裁判年份']
    List = ['当事人犯罪年龄']
    rate = -1
    for i, j in enumerate(loc1):
        if j == '.' or loc2[i] == '.':
            List.append('.')
            continue
        if isinstance(j,int):
            if isinstance(loc2[i],int) or loc2[i].isdigit():
                rate = int(loc2[i]) - j
        elif j.isdigit():
            if isinstance(loc2[i],int) or loc2[i].isdigit():
                rate = int(loc2[i]) - int(j)
        if rate < 0:
            List.append('.')
        else:
            List.append(rate)
    write_to_column(List,1)

def task8_func():
    data = pd.read_excel(excel_file)
    loc = data['文化程度']
    EducationallevelList = [0,1,2,3,4,'0','1','2','3','4']
    data = ['文化程度(过滤)']
    for i in loc:
        if i in EducationallevelList:
            data.append(i)
        else:
            data.append('.')
    write_to_column(data,1)


def task9_func(ColumnName):
    data = pd.read_excel(excel_file)
    loc = data[ColumnName]
    data_list = []
    final_data = [ColumnName+'(过滤)']
    for i in range(0, loc.count()):
        numbers = re.findall(r'[012]', str(loc[i]))
        if numbers == []:
            numbers = ['.']
        data_list.append(set(numbers))
    for Set in data_list:
        tmp_str = '、'.join(Set)
        final_data.append(tmp_str)
    write_to_column(final_data,1)

def task10_func():
    data = pd.read_excel(excel_file)
    loc = data['监禁时间']
    AnList = ['监禁时间(天)']
    for i in range(0, int(loc.count())):
        res = ChangeToDays(loc[i])
        AnList.append(res)
    write_to_column(AnList,1)





if __name__ == '__main__':
    task1_func()
    task2_func()
    task3_func()
    task4_func()
    task5_func()
    task6_func()
    task7_func()
    task8_func()
    task9_func('本次  罪罪名')
    task9_func('曾  罪类型')
    task10_func()